import json
import csv
from decimal import Decimal
from openpyxl import load_workbook
import sqlite3

#from os import walk
#course_details_filelist = ['./data/UG/'+y for y in [x[2] for x in walk('./data/UG')][0]]
#print(course_details_filelist)

no_cap_list = ['ME', 'CS', 'MTech', 'PhD', 'MDS', 'CY', 'II', 'III', 'in', 'and', '2d', 'for', 'a', 'HT', 'MT', 
        'is', 'of', 'to', 'the', 'PH', 'DSP', 'EE', 'LA', 'CA', 'FEM', 'CFD', 'IC',
        'LTE-4G', 'MAC', 'AI', 'ML', 'GIS', '-I', '-II', 'MA', 'BM', 'BO', 'LA/CA', 
        'CMOS', 'AC', 'DC', 'MOS', 'VLSI', 'AdS', 'CFT', 'MHD', 'FES', 'ACM', 'MAD', 'IDM' ]
def capitals(s):
    """capitalize the course titles, skip which are in the no_cap_list."""
    res = ''
    for w in s.split(' '):
        if w not in no_cap_list: w = w.capitalize()
        res += w + ' '
    #print(res)
    return res

import re
tex_conv = {
        '&': r'\&',
        '%': r'\%',
        '$': r'\$',
        '#': r'\#',
        '_': r'\_',
        '{': r'\{',
        '}': r'\}',
        '~': r'\textasciitilde{}',
        '^': r'\^{}',
        '\\?P<symb>[a-zA-Z]': '\\{\\g<symb>}',
        '<': r'\textless{}',
        '>': r'\textgreater{}',
        }
tex_regex = re.compile('|'.join(re.escape(key) for key in sorted(tex_conv.keys(), key = lambda item: - len(item))))
def tex_escape(text):
    """
        :param text: a plain text message
        :return: the message escaped to appear correctly in LaTeX
    """
    return tex_regex.sub(lambda match: tex_conv[match.group()], text)

def sanitize(code, name, credits, segments, pre_req, syl):
    """standard formats for items appearing in output"""
    if code is not None: code = code.replace(' ', '').upper().strip()
    if name is not None: name = capitals(name).replace('&', 'and').strip()
    try: credits = float(credits)
    except: credits = 1.0
    segments = str(segments).strip()
    if segments == None or segments == 'None': segments = ''
    if pre_req == None or pre_req == 'None': pre_req = '' # pre_reqs are not processed as codes, should be.
    else: pre_req = pre_req.replace('&', 'and')
    if syl is not None: syl = syl.replace('&', 'and').replace('\\\\ [', '[') #tex_escape(tex_escape(syl))
    return code, name, credits, segments, pre_req, syl

def get_course_db(level):
    con = sqlite3.connect('./courses_'+level+'.db')
    cur = con.cursor()
    return cur
    #check if it is in the db already
    #if dept=='EP': dept = 'PH'
    #chk = """SELECT name FROM sqlite_master WHERE type='table' AND name='%s_courses';""" % dept
    #chk_res = cur.execute(chk).fetchall()
    #if len(chk_res)==0: 
    #    print("Table %s_courses not found, please upload..." % dept)
    #    sys.exit()
    #return cur

def update_dept_cdesc(dept, sheet, level):
    """Update the department course db at level. Use like this: 
    update_dept_cdesc('EE', wb.get_sheet_by_name("course-descriptions"), 'ug').
    First checks if dept sheet exist in db, if not then create it.
    Then courses are read from excel file (EE_Course_Descriptions.xlsx) and 
    inserted in either EE or ID tables. Course is overwritten if already
    exists."""
    def chk_dept(cur, dept):
        chk = """SELECT name FROM sqlite_master WHERE type='table' AND name='%s_courses';""" % dept
        chk_res = cur.execute(chk).fetchall()
        if len(chk_res)==0: 
            # if there is no table, then create it.
            print("Table %s_courses not found, creating..." % dept)
            # Code is unique, name is not, as same name can be offered for UG and PG.
            c = """
            CREATE TABLE %s_courses (
            code           STRING  PRIMARY KEY UNIQUE NOT NULL,
            name           STRING  NOT NULL,
            credits        float NOT NULL,
            semester       INTEGER,
            pre_req        STRING,
            syllabus       TEXT,
            segments       STRING,
            remarks        TEXT,
            global_remarks TEXT
            );""" % dept
            cur.execute(c)

    con = sqlite3.connect('./courses_'+level.upper()+'.db')
    cur = con.cursor()
    #check if it is in the db already
    chk_dept(cur, dept)
    chk_dept(cur, 'ID')
    con.commit()
    
    d_ins = """INSERT OR REPLACE INTO %s_courses (semester, code, name, credits, segments, pre_req, 
            syllabus, remarks, global_remarks) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""" % dept
    id_ins = """INSERT OR REPLACE INTO %s_courses (semester, code, name, credits, segments, pre_req, 
            syllabus, remarks, global_remarks) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""" % 'ID'
    for r in sheet.iter_rows(min_row=2):# first row contains headings.
        try:
            sem, c, n, cd, seg, pre, syl, rem, g_rem = [r[i].value for i in range(9)] 
            ins_str = id_ins if c[:2]=='ID' else d_ins
            c, n, cd, seg, pre, syl = sanitize(c, n, cd, seg, pre, syl)
            cur.execute(ins_str, [sem, c, n, cd, seg, pre, syl, rem, g_rem])
        except Exception as e: 
            print(e, str(r[1].value), r[1])
            if str(r[1].value)=='None': break # stop when a blank course code encountered.
    con.commit()
    con.close()

def gen_course_description(dept, level, title):
    """course syllabus for dept at level, see describe and syllabus macros in pre-doc.tex"""
    #pre = open("./parts/pre_descr_table.tex").readlines()
    #post = open("./parts/post_descr_table.tex").readlines()
    #print("".join(pre))
    if dept.upper()=='ES': return
    cur = get_course_db(level)
    print("\\section{%s}"%title)
    #print("""\\newpage""")
    #print("""\\textbf{%s} %s Course Description""" % (dept, level.upper()))
    table_pre = "\\columnratio{0.25} \\setlength{\\columnsep}{1em} \\begin{paracol}{2}"
    print(table_pre)
    row_style = "\\describe{%s}{%s}{%s}{%s} \\switchcolumn \\vspace{0mm} \\syllabus{%s} \\switchcolumn[0]*"
    q = """SELECT code, name, credits, segments, pre_req, syllabus FROM %s_courses ORDER BY code""" % dept
    rows = cur.execute(q).fetchall()
    for row in rows: 
        code, name, credits, segments, pre_req, syl = row
        code, name, credits, segments, pre_req, syl = sanitize(code, name, credits, segments, pre_req, syl)
        credits = Decimal(credits)
        if code[2:] == 'XXXX': continue
        if pre_req != None and pre_req != '': pre_req = "\\triangleright "+pre_req
        print(row_style % (code, credits, name, pre_req, syl))
    table_post = "\\end{paracol}"
    print(table_post)
    #print("".join(post))

seg_dict = {}
def get_segment_line(seg):
    """return latex code for a rounded bar with segment written inside it. The length
    and starting position of bar is as per segment specification"""
    width = 6
    try: return seg_dict[seg]
    except: 
        template = "{\\hspace{%d mm} \\begin{tcolorbox}[colback=lightblue,boxsep=0.3mm,top=0pt,bottom=0pt,width=%d mm,boxrule=0.1mm]{\\begin{center}{\\textbf{\\tiny{%s}}}\\end{center}}\\end{tcolorbox} \\hspace{%d mm}}"
        if seg != None and seg != '':
            try: start, end = [float(x) for x in seg]
            except: return seg_dict['blank']
            frontspace = (start - 1)*width
            endspace = (width - end)*width
            midspace = (end-start+1)*width
            tbox = template % (frontspace, midspace, '%s', endspace)
            seg_dict[seg] = tbox
            return tbox
        else:
            try: return seg_dict['blank']
            except:
                seg_dict['blank'] = "{\\tiny{%s}}"
                return seg_dict['blank']

def gen_curriculum(dept, sheet, title, level, display_seg=True):
    """curriculum for dept at level using a given sheet, with title taken from sheet name.
    displays a segment column if asked. 
    See currformat and surrformat macros for single lines with and w/o display-seg in pre-doc.tex.
    Course credits range is not supported at the moment."""
    #pre = open("./parts/pre_curr_table.tex").readlines()
    #post = open("./parts/post_curr_table.tex").readlines()
    #print("".join(pre))
    cur = get_course_db(level)
    #table_pre = "\\begin{longtable}{@{}llll@{}} \
    #\\textcolor{Grey}{\\textbf{Course Code}} & \\textcolor{Grey}{\\textbf{Course Name}} & \
    #\\textcolor{Grey}{\\textbf{Credits}} & \\textcolor{Grey}{\\textbf{Segments}} \\\\ \
    #\\toprule \\endhead"
    #{@{}rlll@{}} \
    #table_pre = """\\setlength\\LTleft{0pt} \\setlength\\LTright{0pt} \\begin{longtable}{@{\extracolsep{\\fill}}rlll@{}}
    
    #print("\\hspace{-1cm}")
    #print("""\\vspace{1cm}""")
    print("""\\section{%s}"""%(title))
    #table_pre = """\\begin{longtable}{@{\\extracolsep{\\fill}}rllr@{}} \
    table_pre = """\\setlength\\LTleft{0pt} \\setlength\\LTright{0pt} \n \\begin{longtable}{@{\\extracolsep{\\fill}}rllr@{}} \
            \\textcolor{Grey}{\\textbf{Cred.}} & \\textcolor{Grey}{\\textbf{Code}} & \
    \\textcolor{Grey}{\\textbf{Course Title}} & \\textcolor{Grey}{\\textbf{Segments}} \\\\ \
    \\toprule \\endhead"""
    table_pre1 = """\\setlength\\LTleft{0pt} \\setlength\\LTright{0pt} \n \\begin{longtable}{@{\\extracolsep{\\fill}}rll@{}} \
            \\textcolor{Grey}{\\textbf{Cred.}} & \\textcolor{Grey}{\\textbf{Code}} & \
    \\textcolor{Grey}{\\textbf{Course Title}} \\\\ \
    \\toprule \\endhead"""
    if display_seg: print(table_pre)
    else: print(table_pre1)
    row_style = "\\currformat{%s}{%s}{%s}"
    #row_style = "\\currformat{%s}{%s}{%s}{\\{black}{lightgreen}{\\tiny{\\hspace{1cm}%s\\hspace{1cm}}}}"
    #row_style = "\\currformat{%s}{%s}{%s}{\\colorbox{lightgreen}{\\framebox(100,2){\\tiny{\\hspace{1cm}%s\\hspace{1cm}}}}}"
    row_style1 = "\\surrformat{%s}{%s}{%s}"

    q = """SELECT name, credits, segments FROM %s_courses WHERE code='%s'"""
    g_remarks = []
    remarks, rem_tex = [], ''
    sem, t1, t2 = None, Decimal(0), Decimal(0)
    # t1 = total credits, t2 = credits for a semester, gt=total at the end
    sem_str = sheet['A1'].value
    b = 0 # base
    if sem_str[:3].lower()=='bas': b=1
    for r in sheet.iter_rows(min_row=2):
        s = r[0].value
        if sem != s:
            if sem != None:
                if display_seg:
                    if b==0:print("\\textbf{%s} & Total & \\\\"%int(t2))
                    print(" & & & \\\\")
                else:
                    if b==0:print("\\textbf{%s} & Total \\\\"%int(t2))
                    print(" & & \\\\")
            t1 += t2
            sem = s
            sem1 = None
            if type(s)==type(1.0): sem1 = str(int(s))
            else: sem1 = s
            t2 = Decimal(0)
            if sem != None: 
                if b==1:# if baskets are there, don't print sem_str
                    if display_seg: print(" \\multicolumn{2}{l}{\\textbf{%s}} & & \\\\" % (sem1))
                    else: print(" \\multicolumn{2}{l}{\\textbf{%s}} & \\\\" % (sem1))
                else:
                    if display_seg: print(" \\multicolumn{2}{l}{\\textbf{%s %s}} & & \\\\" % (sem_str, sem1))
                    else: print(" \\multicolumn{2}{l}{\\textbf{%s %s}} & \\\\" % (sem_str, sem1))
        rem_tex = ''
        if str(r[b+5].value) != 'None':
            matched = False
            i = 0
            for i in range(len(remarks)):
                if remarks[i]==str(r[b+5].value): 
                    matched = True
                    break
            if matched: rem_tex = "$^%d$"%(i+1)
            else:
                remarks.append(str(r[b+5].value))
                rem_tex = "$^%d$"%(len(remarks))
            #print(remarks, m, rem_tex)
        code = r[b+1].value
        if str(r[b+1].value)=='None': break # stop when a blank course code encountered.
        dept = code[:2].upper()
        #print(code, dept)
        if code[2:] == 'XXXX': name, credits, segments = r[b+2].value, r[b+3].value, r[b+4].value
        else: 
            try: name, credits, segments = cur.execute(q % (dept, code)).fetchall()[0]
            except: name, credits, segments = r[b+2].value, r[b+3].value, r[b+4].value
        #print(code, name, credits, segments)
        code, name, credits, segments, pre_req, syl = sanitize(code, name, credits, segments, '', '')
        credits = Decimal(credits)
        #segments = str(segments).strip()
        #if segments == None or segments == 'None': segments = ''
        t1 += credits
        t2 += credits
        #print(row_style % (code, name.title(), credits, segments))
        if display_seg: 
            tbox = get_segment_line(segments)
            print((row_style+tbox) % (credits, code+rem_tex, name, segments))
        else: print(row_style1 % (credits, code+rem_tex, name))
        if r[b+6].value != None: g_remarks.append(r[b+6].value)
    #if sem_str[:2].lower()!='ba' and t2 != 0: print("\\textbf{%s} & Total & \\\\"%t2)
    #if sem_str[:2].lower()!='ba': print("\\hline \\textbf{%d} & Grand Total & \\\\ \\end{longtable}"%t1)
    if b==0 and t2 != 0: print("\\textbf{%s} & Total & \\\\"%t2)
    if b==0: print("\\hline \\textbf{%d} & Grand Total & \\\\ \\end{longtable}"%t1)
    else: print("\\hline \\\\ \\end{longtable}")
    if len(remarks) != 0:
        print("%remarks\n\\begin{footnotesize} \\begin{enumerate}\n")
        for i in range(len(remarks)):
            print("\\item %s\n" %(remarks[i]))
        print("\\end{enumerate}\n \end{footnotesize} ")
    if len(g_remarks) != 0:
        print("%global remarks\n\\begin{itemize}\n")
        for x in g_remarks: print("\\item %s\n" % x.replace('&', 'and'))
        print("\\end{itemize}\n")
    #print(g_remarks.replace('&', 'and') +" \n")
    print("""\\vspace{6mm}""")
    #print("".join(post))

def print_part(f): print(''.join(open(f).readlines()))

import sys
if __name__ == "__main__":
    def get_deptname(f): return f.split('-')[0].split('/')[-1].upper()

    import proc_list
    def update_dept(dept, level):
        fname = proc_list.basedatadir + dept+'_'+level.upper()+'_CourseDescription.xlsx'
        wb = load_workbook(fname)
        desc = wb.get_sheet_by_name("course-descriptions")
        update_dept_cdesc(dept, desc, level)
    def print_level_curr(dept, level):
        fname = proc_list.basedatadir + dept+'_'+level.upper()+'_Curriculum.xlsx'
        wb = load_workbook(fname)
        for s in wb.sheetnames:
            sheet = wb.get_sheet_by_name(s)
            disp_seg = False
            if s[:4] == 'Bach': disp_seg = True
            gen_curriculum(dept, sheet, capitals(s), level, display_seg=disp_seg)
        #gen_course_description(dept, level)
        
    if len(sys.argv) > 1:
        if sys.argv[1] == "update":
            #updates course database of given department: python3 main.py CS ug
            for d in [sys.argv[2]]: #proc_list.depts
                update_dept(d, sys.argv[3]) 

        elif sys.argv[1] == "print-doc":
            # prints whole doc: python3 main.py print-doc
            print_part('./parts/pre-doc.tex')
            print("\\chapter*{2 B.Tech Course Curriclum}\\stepcounter{chapter}\\addcontentsline{toc}{chapter}{2 B.Tech Course Curriculum}")
            import proc_list
            for s in ['curriculum', 'baskets', 'minor', 'honors', 'double-major']:
                for f in proc_list.ug_plist:
                    deptname = get_deptname(f[0]) 
                    wb = load_workbook(f[0])
                    if s in f[1]:
                        sheet = wb.get_sheet_by_name(s)
                        gen_curriculum(deptname, sheet, "UG %s Curriculum" % (capitals(s)), display_seg=s[:4]=="curr")
            print("\\chapter*{3 Course Descriptions}\stepcounter{chapter}\\addcontentsline{toc}{chapter}{3 Course Descriptions}")
            for f in proc_list.ug_plist:
                deptname = get_deptname(f[0]) 
                gen_course_description(deptname)
            print_part('./parts/post-doc.tex')

        elif sys.argv[1] == "print-all": # same as above, above is not easy to modify, so.
            print_part('./parts/pre-doc.tex')
            level = sys.argv[2].upper() 
            depts = sorted(['CS','ME']) #, 'AI', 'EE', 'ME', 'CH', 'BO', 'CE', 'CY', 'DS', 'ES', 'LA', 'MA', 'MS', 'PH'])
            dept_prefix = "Department of "
            depts_expand = {'AI':'Artificial Intelligence', 'BM': 'Biomedical Engineering', 'CS': 'Computer Science and Engineering',
                    'EE': 'Electrical Engineering', 'ME': 'Mechanical and Aerospace Engineering', 'CH': 'Chemical Engineering',
                    'BO': 'Biotechnology', 'CE': 'Civil Engineering', 'CY': 'Chemistry', 'DS': 'Design', 
                    'ES': 'Engineering Science', 'LA': 'Liberal Arts', 'MA': 'Mathematics', 'MS': 'Material Science and Metallurgical Engineering',
                    'PH': 'Physics'}
            level_strs = {'UG_curr': "%s Course Curricula - Bachelors",
                    'PG_curr': "%s Course Curricula - Masters and PhD",
                    'UG_desc': "Course Descriptions - Bachelors",
                    'PG_desc': "Course Descriptions - Masters and PhD"}
            def print_level(level, count):
                #dcount = count
                for d in depts:
                    if level == 'UG' and d in ['BM', 'BO', 'CY']: continue
                    #chap_title = level_strs[level+'_curr'] % (dcount, dept_prefix+depts_expand[d])
                    #print("\\chapter*{"+chap_title+"}\\stepcounter{chapter}\\addcontentsline{toc}{chapter}{"+chap_title+"}")
                    chap_title = level_strs[level+'_curr'] % (dept_prefix+depts_expand[d])
                    print("\\chapter{"+chap_title+"}") 
                    print_level_curr(d, level) 
             
                chap_title = level_strs[level+'_desc'] 
                print("\\chapter{"+chap_title+"}") 
                for d in depts:
                    sect_title = dept_prefix+depts_expand[d]
                    gen_course_description(d, level, sect_title)
            print_level(level, 1)
            print_part('./parts/post-doc.tex')
                
        elif sys.argv[1] == "print-one":
            # prints single dept at a level: python3 main.py print-one CS ug
            print_part('./parts/pre-doc.tex')
            dept = sys.argv[2]
            level = sys.argv[3]
            chap_str = "%s %s Curriculum"%(dept, level)
            print("\\chaptertitle{%s}"%chap_str)
            #print("\\chapter*{\\thechapter\\hspace{3mm} %s}\\stepcounter{chapter}\\addcontentsline{toc}{chapter}{\\thechapter\\hspace{3mm} %s}"%(chap_str, chap_str))
            for d in [dept]: #proc_list.depts
                print_level_curr(d, level) 
            print_part('./parts/post-doc.tex')
            
    
    


